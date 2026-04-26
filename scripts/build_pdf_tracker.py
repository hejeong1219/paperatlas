from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/zzeongvely/Library/CloudStorage/Dropbox/llm-wiki")
OUT = ROOT / "outputs" / "pdf-download-tracker-batches-1-7.xlsx"


ROWS = [
    [1, "Chen 2013", "chen-2013-cancer-immunity-cycle", "pending", "", "https://pubmed.ncbi.nlm.nih.gov/23890062/", "PubMed / publisher page", "Immunity", "Publisher or mirror PDF not yet captured", "Ingested from article landing information; local PDF pending."],
    [2, "Chen 2017", "chen-2017-cancer-immune-set-point", "downloaded", str(ROOT / "raw/inbox/papers/chen-2017-cancer-immune-set-point.pdf"), "https://www.nature.com/articles/nature21349", "Nature direct PDF worked", "Nature", "", "Verified local PDF."],
    [3, "Sotillo 2015", "sotillo-2015-cd19-mutations-alternative-splicing-cart19", "pending", "", "https://pubmed.ncbi.nlm.nih.gov/26516065/", "PubMed / Cancer Discovery", "Cancer Discovery", "Publisher or PMC PDF not yet captured", "Needs manual PDF retrieval."],
    [4, "Orlando 2018", "orlando-2018-target-antigen-loss-car19-therapy", "downloaded", str(ROOT / "raw/inbox/papers/orlando-2018-target-antigen-loss-car19-therapy.pdf"), "https://www.nature.com/articles/s41591-018-0146-z", "Nature direct PDF worked", "Nature Medicine", "", "Verified local PDF."],
    [5, "Qiu 2025", "qiu-2025-lineage-switch-cd19-cart-treatment-ball", "downloaded", str(ROOT / "raw/inbox/papers/qiu-2025-lineage-switch-cd19-cart-treatment-ball.pdf"), "https://www.nature.com/articles/s41375-024-02467-2", "Nature/Leukemia PDF worked", "Leukemia", "", "Verified local PDF."],
    [6, "Labanieh 2023", "labanieh-2023-car-immune-cells-design-principles-resistance", "downloaded", str(ROOT / "raw/inbox/papers/labanieh-2023-car-immune-cells-design-principles-resistance.pdf"), "https://www.nature.com/articles/s41586-023-05790-6", "Nature PDF worked", "Nature", "", "Verified local PDF."],
    [7, "Kantarjian 2017", "kantarjian-2017-blinatumomab-versus-chemotherapy-advanced-all", "pending", "", "https://pubmed.ncbi.nlm.nih.gov/28249141/", "PubMed / NEJM / Ovid", "NEJM", "Scripted PDF capture not obtained", "Needs manual PDF retrieval."],
    [8, "Moreau 2022", "moreau-2022-teclistamab-relapsed-refractory-multiple-myeloma", "pending", "", "https://pubmed.ncbi.nlm.nih.gov/35857662/", "PubMed / NEJM / Ovid", "NEJM", "Scripted PDF capture not obtained", "Needs manual PDF retrieval."],
    [9, "Lesokhin 2023", "lesokhin-2023-elranatamab-relapsed-refractory-multiple-myeloma", "downloaded", str(ROOT / "raw/inbox/papers/lesokhin-2023-elranatamab-relapsed-refractory-multiple-myeloma.pdf"), "https://www.nature.com/articles/s41591-023-02528-9", "Nature direct PDF worked", "Nature Medicine", "", "Verified local PDF."],
    [10, "Ahn 2023", "ahn-2023-tarlatamab-previously-treated-small-cell-lung-cancer", "pending", "", "https://www.ovid.com/journals/nejm/fulltext/10.1056/nejmoa2307980~tarlatamab-for-patients-with-previously-treated-small-cell", "Ovid article page", "NEJM", "PDF endpoint not captured in this environment", "Ingested from accessible article page; manual PDF may be easier in browser."],
    [11, "Mountzios 2025", "mountzios-2025-tarlatamab-small-cell-lung-cancer-after-platinum", "pending", "", "https://www.ovid.com/journals/nejm/fulltext/10.1056/nejmoa2502099~tarlatamab-in-small-cell-lung-cancer-after-platinum-based", "Ovid article page / PubMed", "NEJM", "PDF endpoint exists but not captured locally", "PubMed abstract accessible; manual browser PDF may work."],
    [12, "Harding 2023", "harding-2023-zanidatamab-her2-amplified-biliary-tract-cancer", "pending", "", "https://www.sciencedirect.com/science/article/pii/S1470204523002425", "ScienceDirect page; UCL Discovery mirror", "Lancet Oncology", "ScienceDirect PDF blocked by Cloudflare", "Try publisher in browser; UCL Discovery hosts an open repository record."],
    [13, "Pant 2026", "pant-2026-zanidatamab-her2-positive-metastatic-biliary-tract-cancer", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC12635922/", "PMC article page", "JAMA Oncology", "PMC PDF not captured via script", "Full PMC text accessible for manual PDF/browser save."],
    [14, "Goebeler 2024", "goebeler-2024-bispecific-multispecific-antibodies-oncology", "downloaded", str(ROOT / "raw/inbox/papers/goebeler-2024-bispecific-multispecific-antibodies-oncology.pdf"), "https://www.nature.com/articles/s41571-024-00905-y", "Nature direct PDF worked", "Nature Reviews Clinical Oncology", "", "Verified local PDF."],
    [15, "Modi 2022", "modi-2022-trastuzumab-deruxtecan-her2-low-advanced-breast-cancer", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC10561652/", "PMC article page", "NEJM", "PMC PDF URL returns HTML challenge page when scripted", "Full PMC text accessible; manual browser PDF likely easier."],
    [16, "Bardia 2024", "bardia-2024-trastuzumab-deruxtecan-after-endocrine-therapy-metastatic-breast-cancer", "pending", "", "https://www.ovid.com/journals/nejm/fulltext/10.1056/nejmoa2407086~trastuzumab-deruxtecan-after-endocrine-therapy-in-metastatic", "Ovid article page / SNU abstract mirror", "NEJM", "PDF endpoint exists but not captured locally", "Structured abstract available from SNU Pure page for verification."],
    [17, "Chen 2026", "chen-2026-trastuzumab-deruxtecan-resistance-her2-expression-binding", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC12631751/", "PMC article page", "Cancer Discovery", "PMC PDF URL returns HTML challenge page when scripted", "Full PMC text accessible; manual browser PDF likely easier."],
    [18, "Chmielecki 2023", "chmielecki-2023-acquired-resistance-first-line-osimertinib", "downloaded", str(ROOT / "raw/inbox/papers/chmielecki-2023-acquired-resistance-first-line-osimertinib.pdf"), "https://www.nature.com/articles/s41467-023-35961-y", "Nature direct PDF worked", "Nature Communications", "", "Verified local PDF."],
    [19, "Han 2023", "han-2023-tumour-microenvironment-changes-osimertinib-resistance", "pending", "", "https://www.sciencedirect.com/science/article/pii/S0959804923002496", "ScienceDirect open-access article page", "European Journal of Cancer", "Direct pdfft download blocked by Cloudflare in scripted environment", "Article is open access and shows View PDF in search snippets; browser download may work."],
    [20, "Grasso 2018", "grasso-2018-genetic-mechanisms-immune-evasion-colorectal-cancer", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC5984687/", "PMC article page", "Cancer Discovery", "PMC PDF URL returns HTML challenge page when scripted", "Full PMC text accessible; manual browser PDF likely easier."],
    [21, "Zaretsky 2016", "zaretsky-2016-acquired-resistance-pd1-melanoma", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC5007206/", "PMC article page / NEJM", "NEJM", "PMC PDF route remained challenge-gated", "Full PMC text accessible; manual browser PDF likely easier."],
    [22, "Shin 2017", "shin-2017-primary-resistance-pd1-jak-mutations", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC5296316/", "PMC article page / AACR", "Cancer Discovery", "No valid local PDF captured", "Full PMC text accessible and ingested."],
    [23, "Gettinger 2017", "gettinger-2017-hla-antigen-presentation-checkpoint-lung", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC5718941/", "PMC article page", "Cancer Discovery", "PMC PDF endpoint returned HTML challenge page", "Bad placeholder PDF removed from raw; manual browser save likely easier."],
    [24, "McGranahan 2017", "mcgranahan-2017-allele-specific-hla-loss-lung-cancer", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC5720478/", "PMC article page / Cell", "Cell", "PMC PDF endpoint returned HTML challenge page", "Bad placeholder PDF removed from raw; full text still ingested."],
    [25, "Tran 2016", "tran-2016-kras-tcell-transfer-hla-loss", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC5178827/", "PMC article page / NEJM", "NEJM", "PMC PDF endpoint returned HTML challenge page", "Bad placeholder PDF removed from raw; full text still ingested."],
    [26, "Liu 2026", "liu-2026-usp22-ezh2-mhc1-checkpoint-resistance", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC12721901/", "PMC article page / JCI article page", "Journal of Clinical Investigation", "JCI and PMC PDF endpoints returned HTML challenge content", "Bad placeholder PDF removed from raw; full text still ingested."],
    [27, "Zhao 2025", "zhao-2025-interferon-epsilon-9p21-immune-cold-tumors", "pending", "", "https://www.sciencedirect.com/science/article/abs/pii/S1556086424025395", "ScienceDirect abstract page", "Journal of Thoracic Oncology", "No valid local PDF captured", "Ingested from PubMed and publisher abstract page."],
    [28, "Toboni 2023", "toboni-2023-mlh1-hypermethylation-lynch-like-endometrial-cancer", "pending", "", "https://pubmed.ncbi.nlm.nih.gov/37683549/", "PubMed / ScienceDirect", "Gynecologic Oncology", "No valid local PDF captured", "PubMed abstract and publisher highlights used for ingest."],
    [29, "Yang 2021", "yang-2021-msih-gastric-cancer-heterogeneity-worse-survival", "pending", "", "https://jmg.bmj.com/content/58/1/12", "BMJ article page", "Journal of Medical Genetics", "Direct PDF endpoint returned HTML rather than a PDF", "Bad placeholder PDF removed from raw; abstract-level ingest completed."],
    [30, "Mariathasan 2018", "mariathasan-2018-tgfb-pdl1-blockade-tcell-exclusion", "downloaded", str(ROOT / "raw/inbox/papers/mariathasan-2018-tgfb-pdl1-blockade-tcell-exclusion.pdf"), "https://www.nature.com/articles/nature25501", "Nature direct PDF worked", "Nature", "", "Verified local PDF."],
    [31, "Laino 2020", "laino-2020-serum-il6-crp-melanoma-checkpoint-inhibition", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC7312339/", "PMC article page / JITC article page", "Journal for ImmunoTherapy of Cancer", "BMJ PDF route returned HTML rather than a valid PDF", "Full PMC text accessible and ingested; bad placeholder PDF removed from raw."],
    [32, "Elyada 2019", "elyada-2019-apcafs-pancreatic-cancer", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC6727976/", "PMC article page", "Cancer Discovery", "No valid local PDF captured", "Full PMC text accessible and ingested."],
    [33, "Allen 2017", "allen-2017-antiangiogenic-pdl1-hev-formation", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC5554432/", "PMC article page / PubMed", "Science Translational Medicine", "No valid local PDF captured", "Full PMC text accessible and ingested."],
    [34, "Zhu 2014", "zhu-2014-csf1r-blockade-pancreatic-cancer-models", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC4182950/", "PMC article page", "Cancer Research", "PMC PDF endpoint returned HTML challenge content", "Full PMC text accessible and ingested."],
    [35, "Ruella 2023", "ruella-2023-cart-resistance-haematological-malignancies", "downloaded", str(ROOT / "raw/inbox/papers/ruella-2023-cart-resistance-haematological-malignancies.pdf"), "https://www.nature.com/articles/s41573-023-00807-1", "Nature direct PDF worked", "Nature Reviews Drug Discovery", "", "Verified local PDF."],
    [36, "Sterner 2021", "sterner-2021-cart-current-limitations-potential-strategies", "downloaded", str(ROOT / "raw/inbox/papers/sterner-2021-cart-current-limitations-potential-strategies.pdf"), "https://www.nature.com/articles/s41408-021-00459-7", "Nature direct PDF worked", "Blood Cancer Journal", "", "Verified local PDF after replacing an invalid PMC HTML placeholder."],
    [37, "Zugasti 2025", "zugasti-2025-cart-cancer-current-challenges-future-directions", "downloaded", str(ROOT / "raw/inbox/papers/zugasti-2025-cart-cancer-current-challenges-future-directions.pdf"), "https://www.nature.com/articles/s41392-025-02269-w", "Nature direct PDF worked", "Signal Transduction and Targeted Therapy", "", "Verified local PDF."],
    [38, "Helmink 2020", "helmink-2020-b-cells-tls-immunotherapy-response", "downloaded", str(ROOT / "raw/inbox/papers/helmink-2020-b-cells-tertiary-lymphoid-structures-immunotherapy-response.pdf"), "https://www.nature.com/articles/s41586-019-1922-8", "Nature direct PDF worked", "Nature", "", "Verified local PDF."],
    [39, "Italiano 2022", "italiano-2022-pembrolizumab-soft-tissue-sarcomas-tls", "downloaded", str(ROOT / "raw/inbox/papers/italiano-2022-pembrolizumab-soft-tissue-sarcomas-tls.pdf"), "https://www.nature.com/articles/s41591-022-01821-3", "Nature direct PDF worked", "Nature Medicine", "", "Verified local PDF."],
    [40, "Sun 2022", "sun-2022-tls-neoadjuvant-chemoimmunotherapy-resectable-nsclc", "pending", "", "https://jitc.bmj.com/content/10/11/e005531", "JITC article page", "Journal for ImmunoTherapy of Cancer", "BMJ PDF route returned HTML rather than a valid PDF", "Abstract-level ingest completed; bad placeholder PDF removed from raw."],
    [41, "Du 2025", "du-2025-tls-gene-signature-advanced-nsclc", "downloaded", str(ROOT / "raw/inbox/papers/du-2025-tls-gene-signature-advanced-nsclc.pdf"), "https://link.springer.com/article/10.1007/s00262-025-04165-2", "Springer direct PDF worked", "Cancer Immunology, Immunotherapy", "", "Verified local PDF."],
    [42, "Harmon 2020", "harmon-2020-immune-consequences-lactate-tumor-microenvironment", "pending", "", "https://link.springer.com/book/10.1007/978-3-030-43093-1", "Springer book landing page / chapter listing", "Advances in Experimental Medicine and Biology", "Scripted capture hit a 404 HTML page rather than a valid PDF", "Chapter-level ingest completed from listing context; bad placeholder file remains non-PDF and should not be treated as downloaded."],
    [43, "Melero 2025", "melero-2025-gdf15-overcome-pd1-pdl1-resistance", "downloaded", str(ROOT / "raw/inbox/papers/melero-2025-gdf15-overcome-pd1-pdl1-resistance.pdf"), "https://www.nature.com/articles/s41586-024-08305-z", "Nature direct PDF worked", "Nature", "", "Verified local PDF."],
    [44, "Chen 2018", "chen-2018-exosomal-pdl1-immunosuppression-pd1-response", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC6095740/", "PMC full text / Nature article page", "Nature", "PMC PDF route returned HTML challenge page rather than a valid PDF", "Full PMC text accessible and ingested; bad placeholder file remains non-PDF and should not be treated as downloaded."],
    [45, "Simpson 2023", "simpson-2023-gut-microbiota-enhance-checkpoint-efficacy", "downloaded", str(ROOT / "raw/inbox/papers/simpson-2023-gut-microbiota-enhance-checkpoint-efficacy.pdf"), "https://www.nature.com/articles/s41571-023-00803-9", "Nature direct PDF worked", "Nature Reviews Clinical Oncology", "", "Verified local PDF."],
    [46, "Almonte 2026", "almonte-2026-gut-dysbiosis-oncology-immunoresistance", "downloaded", str(ROOT / "raw/inbox/papers/almonte-2026-gut-dysbiosis-oncology-immunoresistance.pdf"), "https://www.nature.com/articles/s41422-025-01212-6", "Nature direct PDF worked", "Cell Research", "", "Verified local PDF."],
    [47, "Skoulidis 2018", "skoulidis-2018-stk11-lkb1-pd1-resistance-kras-lung", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC6030433/", "PMC full text / AACR article page", "Cancer Discovery", "PMC PDF route returned HTML challenge page rather than a valid PDF", "Full PMC text accessible and ingested; bad placeholder file remains non-PDF and should not be treated as downloaded."],
    [48, "Skoulidis 2024", "skoulidis-2024-ctla4-abrogates-keap1-stk11-resistance", "downloaded", str(ROOT / "raw/inbox/papers/skoulidis-2024-ctla4-abrogates-keap1-stk11-resistance.pdf"), "https://www.nature.com/articles/s41586-024-07943-7", "Nature direct PDF worked", "Nature", "", "Verified local PDF."],
    [49, "Johnson 2023", "johnson-2023-poseidon-first-line-metastatic-nsclc", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC9937097/", "PMC full text / PubMed", "Journal of Clinical Oncology", "No valid local PDF captured yet", "Full PMC text accessible and ingested."],
    [50, "Cristescu 2018", "cristescu-2018-pan-tumor-genomic-biomarkers-pd1-blockade", "pending", "", "https://www.science.org/doi/10.1126/science.aar3593", "Science article page / abstract mirrors", "Science", "No valid local PDF captured yet", "Ingested from primary abstract-level article summaries and institutional mirror."],
    [51, "Sequist 2011", "sequist-2011-genotypic-histological-evolution-egfr-resistance", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC3132801/", "PMC full text / Sci Transl Med article page", "Science Translational Medicine", "PMC PDF route returned HTML challenge page rather than a valid PDF", "Full PMC text accessible and ingested; local placeholder file is not a real PDF."],
    [52, "Lou 2016", "lou-2016-emt-inflammatory-signals-multiple-checkpoints-lung", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC4947453/", "PMC full text / PubMed", "Clinical Cancer Research", "PMC PDF route returned HTML challenge page rather than a valid PDF", "Full PMC text accessible and ingested; local placeholder file is not a real PDF."],
    [53, "Xue 2020", "xue-2020-rapid-nonuniform-adaptation-kras-g12c-inhibition", "downloaded", str(ROOT / "raw/inbox/papers/xue-2020-rapid-nonuniform-adaptation-kras-g12c-inhibition.pdf"), "https://www.nature.com/articles/s41586-019-1884-x", "Nature direct PDF worked", "Nature", "", "Verified local PDF."],
    [54, "Manabe 2022", "manabe-2022-remodeling-tumor-microenvironment-kras-g12c-resistance", "pending", "", "https://www.jci.org/articles/view/156891", "JCI article page", "Journal of Clinical Investigation", "Scripted PDF capture returned HTML rather than a valid PDF", "Article abstract accessible and ingested; local placeholder file is not a real PDF."],
    [55, "Tsai 2022", "tsai-2022-idiosyncratic-resistance-kras-g12c-inhibition", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC8843735/", "PMC full text / JCI article page", "Journal of Clinical Investigation", "PMC PDF route returned HTML challenge page rather than a valid PDF", "Full PMC text accessible and ingested; local placeholder file is not a real PDF."],
    [56, "Oh 2024", "oh-2024-tdxd-cgas-sting-gastric-cancer", "downloaded", str(ROOT / "raw/inbox/papers/oh-2024-tdxd-cgas-sting-gastric-cancer.pdf"), "https://link.springer.com/article/10.1186/s12964-024-01893-3", "Springer direct PDF worked", "Cell Communication and Signaling", "", "Verified local PDF."],
    [57, "Zhang 2021", "zhang-2021-met-amplification-inhibits-sting-immunotherapy", "pending", "", "https://pubmed.ncbi.nlm.nih.gov/34099454/", "PubMed abstract / AACR article page", "Cancer Discovery", "No valid local PDF captured yet", "Ingested from abstract-level article summaries."],
    [58, "Ricciuti 2022", "ricciuti-2022-kras-g12d-nsclc-correlates", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC11006449/", "PMC full text / PubMed", "Annals of Oncology", "No valid local PDF captured yet", "Full PMC text accessible and ingested."],
    [59, "Offin 2019", "offin-2019-tumor-mutation-burden-egfr-tki-efficacy", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC6347551/", "PMC full text / PubMed", "Clinical Cancer Research", "PMC PDF route returned HTML challenge page rather than a valid PDF", "Full PMC text accessible and ingested; local placeholder file is not a real PDF."],
    [60, "Lai 2022", "lai-2022-nivolumab-versus-nivolumab-ipilimumab-egfr-mutant-nsclc", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC9679031/", "PMC full text / PubMed", "JTO Clinical and Research Reports", "PMC PDF route returned HTML challenge page rather than a valid PDF", "Full PMC text accessible and ingested; local placeholder file is not a real PDF."],
    [61, "Nogami 2022", "nogami-2022-impower150-egfr-liver-brain-subgroups", "pending", "", "https://pubmed.ncbi.nlm.nih.gov/34626838/", "PubMed abstract / JTO article page", "Journal of Thoracic Oncology", "No valid local PDF captured yet", "Ingested from abstract-level article summaries."],
    [62, "Wu 2025", "wu-2025-atezolizumab-bevacizumab-pemetrexed-platinum-egfr-tki-failure", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC11666332/", "PMC full text / PubMed", "Clinical and Translational Medicine", "PMC PDF route returned HTML challenge page rather than a valid PDF", "Full PMC text accessible and ingested; local placeholder file is not a real PDF."],
    [63, "Toy 2013", "toy-2013-esr1-ligand-binding-domain-mutations-breast", "downloaded", str(ROOT / "raw/inbox/papers/toy-2013-esr1-ligand-binding-domain-mutations-breast.pdf"), "https://www.nature.com/articles/ng.2822", "Nature direct PDF worked", "Nature Genetics", "", "Verified local PDF."],
    [64, "Robinson 2013", "robinson-2013-activating-esr1-mutations-metastatic-breast", "downloaded", str(ROOT / "raw/inbox/papers/robinson-2013-activating-esr1-mutations-metastatic-breast.pdf"), "https://www.nature.com/articles/ng.2823", "Nature direct PDF worked", "Nature Genetics", "", "Verified local PDF."],
    [65, "Bidard 2022", "bidard-2022-pada1-esr1-switch-fulvestrant-palbociclib", "pending", "", "https://pubmed.ncbi.nlm.nih.gov/36183733/", "PubMed abstract / Lancet Oncology", "Lancet Oncology", "No valid local PDF captured yet", "Ingested from abstract-level trial summaries."],
    [66, "Lopez 2025", "lopez-2025-esr1-mutations-tcell-surveillance-disruption", "downloaded", str(ROOT / "raw/inbox/papers/lopez-2025-esr1-mutations-tcell-surveillance-disruption.pdf"), "https://link.springer.com/article/10.1186/s13058-025-01962-6", "Springer direct PDF worked", "Breast Cancer Research", "", "Verified local PDF."],
    [67, "Lee 2024", "lee-2024-cdk46-senescence-immunogenic-properties", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC10766199/", "PMC full text / PubMed", "Molecular Oncology", "Local scripted capture returned HTML rather than a valid PDF", "Full PMC text accessible and ingested; local placeholder file is not a real PDF."],
    [68, "Wagner 2020", "wagner-2020-senescence-therapeutically-relevant-cdk46", "downloaded", str(ROOT / "raw/inbox/papers/wagner-2020-senescence-therapeutically-relevant-cdk46.pdf"), "https://www.nature.com/articles/s41388-020-1354-9", "Nature direct PDF worked", "Oncogene", "", "Verified local PDF."],
    [69, "Robinson 2015", "robinson-2015-integrative-clinical-genomics-advanced-prostate", "pending", "", "https://pmc.ncbi.nlm.nih.gov/articles/PMC4484602/", "PMC full text / PubMed", "Cell", "Local scripted capture returned HTML rather than a valid PDF", "Full PMC text accessible and ingested; local placeholder file is not a real PDF."],
    [70, "Balbas 2013", "balbas-2013-overcoming-resistance-antiandrogens-rational-design", "downloaded", str(ROOT / "raw/inbox/papers/balbas-2013-overcoming-resistance-antiandrogens-rational-design.pdf"), "https://elifesciences.org/articles/00499", "eLife direct PDF worked", "eLife", "", "Verified local PDF."],
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "pdf-status"

    headers = [
        "ref_no",
        "citation_short",
        "basename",
        "status",
        "local_pdf",
        "article_url",
        "manual_pdf_hint",
        "source_site",
        "block_reason",
        "notes",
    ]
    ws.append(headers)
    for row in ROWS:
        ws.append(row)

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 8,
        "B": 20,
        "C": 52,
        "D": 14,
        "E": 70,
        "F": 72,
        "G": 36,
        "H": 24,
        "I": 42,
        "J": 48,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)


if __name__ == "__main__":
    main()
